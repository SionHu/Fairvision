from csgame.storage_backends import mturk

from django import forms
from django.conf import settings
from django.contrib import admin, messages
from django.contrib.auth import get_user_model
from django.contrib.auth.admin import UserAdmin
from django.contrib.sessions.models import Session
from django.core.files.storage import default_storage
from django.db import connection, transaction
from django.db.migrations.loader import MigrationLoader
from django.db.models import F, Q, Func, Sum, Count, FloatField, Model
from django.db.models.functions import Cast
from django.http import HttpResponse
from django.shortcuts import render
from django.template.defaultfilters import filesizeformat
from django.templatetags.static import static
from django.urls import reverse
from django.utils.html import format_html

from .fields import ListTextInput
from .forms import CustomUserCreationForm, CustomUserChangeForm
from .models import CustomUser, ImageModel, Attribute, Phase, Phase01_instruction, Phase02_instruction, Phase03_instruction, TextInstruction, Answer, Question, HIT


from mturk_hit import create_hit, hitDescriptions

from ast import literal_eval
import base64
import csv
from datetime import datetime
import itertools
from more_itertools import first, partition
import operator
from natsort import natsorted
import tempfile
import zipfile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.writer.excel import save_virtual_workbook

####
# FUNCTIONS INCLUDED FOR BACKWARD COMPATIBILITY
if hasattr(datetime, 'fromisoformat'):
    # added in Python 3.7
    fromisoformat = datetime.fromisoformat
else:
    def fromisoformat(dt):
        return datetime.strptime(dt, "%Y-%m-%dT%H:%M:%S.%f")

try:
    # added in Django 2.2
    from django.db.models.functions import Abs
except:
    from django.db.models.lookups import Transform
    class Abs(Transform):
        function = 'ABS'
        lookup_name = 'abs'

try:
    # replace Django's JSON widget with a better one
    from jsoneditor.forms import JSONEditor
    from django.contrib.postgres.forms import JSONField
    JSONField.widget = JSONEditor
except:
    JSONEditor = forms.Textarea

from django.db.models.manager import BaseManager
def iter_patch(self):
    return iter(self.all())
BaseManager.__iter__ = iter_patch
#
####

def sort_uniq(sequence):
    return map(operator.itemgetter(0), itertools.groupby(natsorted(sequence)))

class CustomUserAdmin(UserAdmin):
    model = CustomUser
    add_form = CustomUserCreationForm
    form = CustomUserChangeForm

def getFolderChoices():
    #calculate the options for the folder
    setChoices = []
    objChoices = []
    for dataset in default_storage.listdir('')[0]:
        if dataset is not 'unknown':
            setChoices.append(dataset)
            objChoices.extend(object for object in default_storage.listdir(dataset)[0] if object is not 'unknown')
    return natsorted(setChoices), sort_uniq(objChoices)

class ImageModelForm(forms.ModelForm):
    img = forms.ImageField(label='Image', widget=forms.FileInput(attrs={'multiple': True}), help_text=(f'Images to upload to S3 ({filesizeformat(settings.DATA_UPLOAD_MAX_MEMORY_SIZE)} or less in total). We only allow JPG files.'), required=True)
    set = forms.CharField(required=True)
    object = forms.CharField(required=True)

    class Meta:
        Model = ImageModel
        fields = ('img', )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if self.instance.pk:
            self.fields['set'].widget = forms.HiddenInput()
            self.fields['object'].widget = forms.HiddenInput()
            self.fields['set'].required = False
            self.fields['object'].required = False

            # get image preview, works now
            self.url = default_storage.url(self.instance.img.name)
        else:

            setChoices, objChoices = getFolderChoices()
            # auto suggestions that pop up when typing in input text field of set or object
            self.fields['set'].widget=ListTextInput(setChoices, 'set')
            self.fields['object'].widget=ListTextInput(objChoices, 'object')
            self.url = None

    def save(self, *args, **kwargs):
        if self.fields['set'].required:
            # multiple file upload
            # NB: does not respect 'commit' kwarg
            with transaction.atomic(), default_storage.upload_lock(self.cleaned_data['set'], self.cleaned_data['object']):
                file_list = natsorted(self.files.getlist('img'.format(self.prefix)), key=operator.attrgetter('name'))
                self.instance.img = file_list[0]
                output = super().save(*args, **kwargs)
                # save the rest of the images to the instances
                ImageModel.objects.bulk_create([
                    ImageModel(img=file) for file in file_list[1:]
                ])
            return output
        else:
            return super().save(*args, **kwargs)


# Filter for dataset folder
class ImageModelDatasetListFilter(admin.SimpleListFilter):
    title = 'dataset'
    parameter_name = 'dataset'
    def lookups(self, request, model_admin):
        ''' Get list of all datasets in the database '''
        return [(name, name) for name in getFolderChoices()[0]]
    def queryset(self, request, queryset):
        val = self.value()
        return queryset.filter(img__contains=val+'/') if val else queryset

# Filter for object folder
class ImageModelObjectListFilter(admin.SimpleListFilter):
    title = 'object'
    parameter_name = 'object'
    def lookups(self, request, model_admin):
        ''' Get list of all object types in the database '''
        return [(name, name) for name in getFolderChoices()[1]]
    def queryset(self, request, queryset):
        val = self.value()
        return queryset.filter(img__contains='/'+val+'/') if val else queryset

class ImageModelAdmin(admin.ModelAdmin):
    form = ImageModelForm
    fields = ('img', 'set', 'object', 'cluster')
    list_display = ('img', 'cluster')
    list_display_links = ('img', 'cluster')
    list_filter = (ImageModelDatasetListFilter, ImageModelObjectListFilter)

    def get_readonly_fields(self, request, obj=None):
        return [] if obj is None else ['img']

class ExportCSVMixin:
    def export_csv(self, request, queryset):
        # https://docs.djangoproject.com/en/1.11/howto/outputting-csv/
        # https://stackoverflow.com/questions/18685223/how-to-export-django-model-data-into-csv-file

        # setup csv writer
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment;filename='+self.export_filename
        writer = csv.writer(response)

        writer.writerow(self.export_field_names)

        # output data
        for obj in queryset:
            writer.writerow([getattr(self, field)(obj) if hasattr(self, field) else getattr(obj, field) for field in self.export_field_names])
        return response
    export_csv.short_description = "Export selected %(verbose_name_plural)s as csv"

class Round2(Func):
    # https://stackoverflow.com/a/55905983
    function = 'ROUND'
    arity = 2
    arg_joiner = '::numeric, '

    def __init__(self, number, places=2, *kwargs, **extra):
        super().__init__(number, places, *kwargs, **extra)

    def as_sqlite(self, compiler, connection, **extra_context):
        return super().as_sqlite(compiler, connection, arg_joiner=", ", **extra_context)

class AttributeAdmin(admin.ModelAdmin, ExportCSVMixin):
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.annotate(_weight=Round2(Cast(F('answer__count'), FloatField()) / Cast(Count('answer__question__answers'), FloatField())))
    #    sum = Attribute.objects.all().aggregate(Sum('count'))['count__sum']
    #    return qs.annotate(bias=Round2(Abs(Cast(F('count'), FloatField()) / float(sum)), 2))

    #def bias(self, obj):
    #    return f"{obj.bias:.2f}"
    def weight(self, obj):
        return f"{obj._weight:.2f}"
    weight.admin_order_field = "_weight"
    def Answer(self, obj):
        return format_html("<a href={}>{}</a>".format(
            reverse('admin:{}_{}_change'.format(Answer._meta.app_label, Answer._meta.model_name),
            args=(obj.answer.pk,)),
        obj.answer))
    def question(self, obj):
        return format_html("<a href={}>{}</a>".format(
            reverse('admin:{}_{}_change'.format(Question._meta.app_label, Question._meta.model_name),
            args=(obj.answer.question.pk,)),
        obj.answer.question))
    def get_fields(self, request, obj=None):
        if obj: # editing an existing object
            return self.fields + ('question', 'Answer')
        return self.fields + ('answer',)
    def get_readonly_fields(self, request, obj=None):
        if obj: # editing an existing object
            return self.readonly_fields + ('question', 'Answer')
        return self.readonly_fields
    fields = ('word', 'count')
    list_display = ('word', 'count', 'weight')
    export_filename = 'phase3-attributes.csv'
    export_field_names = ['word','count','weight']
    actions = ['export_csv']

class PhaseForm(forms.ModelForm):
    class Meta:
        model = Phase
        fields = '__all__'
        widgets = {
            'get': JSONEditor,
            'post': JSONEditor
        }
    def save(self, *args, **kwargs):
        obj = self.instance
        obj.get = literal_eval(self.cleaned_data['get'])
        obj.post = literal_eval(self.cleaned_data['post'])
        return super().save(*args, **kwargs)

class PhaseAdmin(admin.ModelAdmin):
#    form = PhaseForm
    readonly_fields = ('phase',)
    fieldsets = (
        (None, {'fields': ('phase', 'get', 'post', 'imgset')}),
    )
    def has_add_permission(self, request):
        return False
    def has_change_permission(self, request, obj=None):
        return True

class SessionForm(forms.ModelForm):
    decoded_data = forms.CharField(widget=JSONEditor)
    class Meta:
        model = Session
        fields = ('session_key', 'decoded_data', 'expire_date')
        exclude = ("session_data", )
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        public, self.private = partition(lambda k: k[0].startswith('_'), self.instance.get_decoded().items())
        self.initial['decoded_data'] = dict(public) if self.instance else '{}'
        self.objects = self._meta.model.objects
    def save(self, *args, **kwargs):
        obj = self.instance
        data = literal_eval(self.cleaned_data['decoded_data'])
        data.update(self.private)
        obj.session_data = self.objects.encode(data)
        return super().save(*args, **kwargs)

class SessionAdmin(admin.ModelAdmin):
    form = SessionForm
    list_display = ('session_key', 'username', 'expire_date')
    readonly_fields = ('session_key', 'username')
    fieldsets = (
        (None, {'fields': ('session_key', 'username', 'decoded_data', 'expire_date')}),
    )
    def has_add_permission(self, request):
        return False
    def username(self, obj):
        id = obj.get_decoded().get('_auth_user_id')
        if id is None: return None
        user = CustomUser.objects.get(id=id)
        #return .username if id else None
        return format_html("<a href={}>{}</a>".format(
            reverse('admin:{}_{}_change'.format(user._meta.app_label, user._meta.model_name),
            args=(user.pk,)),
        user.username))

class HITApprovalForm(forms.Form):
    reason = forms.CharField(label='Reason for Acceptance', max_length=256, required=False, widget=forms.Textarea(attrs={'cols': 80, 'rows': 20}))
    bonus = forms.DecimalField(label='Worker Bonus', max_value=1, min_value=0, decimal_places=2, initial='0.00', widget=forms.NumberInput(attrs={'step': 0.01}))

    def clean(self):
        cleaned_data = super().clean()
        reason = cleaned_data.get("reason")
        bonus = cleaned_data.get("bonus")

        if bonus and not reason:
            raise forms.ValidationError(
                "If bonus is specified, than a reason is also required."
            )
        return cleaned_data

class HITRejectionForm(forms.Form):
    reason = forms.CharField(label='Reason for Rejection (required)', max_length=256, widget=forms.Textarea(attrs={'cols': 80, 'rows': 20, 'maxlength': 256}))

class HITCreationForm(forms.ModelForm):
    PHASE_CHOICES = [
        ('phase01a', 'Phase 1a'),
        ('phase01b', 'Phase 1b'),
        ('phase03', 'Phase 3'),
    ]
    assignment_id = forms.CharField(widget=forms.HiddenInput(), initial='ASSIGNMENT_ID_NOT_AVAILABLE')
    data = forms.CharField(widget=forms.HiddenInput(), initial="{'_':''}")
    phase = forms.ChoiceField(label='Phase', choices=PHASE_CHOICES, initial='phase01a')
    count = forms.DecimalField(label='Number of HITs', min_value=1, decimal_places=0, initial='1')

    class Meta:
        model = HIT
        exclude = ()

    def __init__(self, *args, **kwargs):
        HIT.objects.filter(assignment_id='ASSIGNMENT_ID_NOT_AVAILABLE').delete()
        super().__init__(*args, **kwargs)

    def save(self, *args, **kwargs):
        #create_hit(self.cleaned_data['phase'], self.cleaned_data['count'])
        print("did it")
        return super().save(*args, **kwargs)

class HITStatusFilter(admin.SimpleListFilter):
    title = 'status'
    parameter_name = 'status'
    def lookups(self, request, model_admin):
        return [('*', 'All'), ('None', 'None'), ('Submitted', 'Submitted'), ('Approved', 'Approved'), ('Rejected', 'Rejected')]
    def queryset(self, request, queryset):
        val = self.value()
        if val == '*':
            return queryset
        if val == 'None':
            return queryset.filter(~Q(data__has_key='status') | Q(data__status=''))
        return queryset.filter(data__status=val)
    def value(self):
        value = super().value()
        return '*' if value is None else value
        return 'Submitted' if value is None else value
    def choices(self, changelist):
        return ({
            'selected': self.value() == str(lookup),
            'query_string': changelist.get_query_string({self.parameter_name: lookup}),
            'display': title,
        } for lookup, title in self.lookup_choices)

class HITWorkerFilter(admin.SimpleListFilter):
    title = 'Worker ID'
    parameter_name = 'workerID'
    def lookups(self, request, model_admin):
        return []
    def queryset(self, request, queryset):
        val = self.value()
        if val is None:
            return queryset
        return queryset.filter(data__workerID=val)

class HITAdmin(admin.ModelAdmin):
    list_display = ['assignment_id', 'status', 'questions', 'worker_id', 'work_time', 'phase', 'start_time']
    readonly_fields = ('assignment_id', 'hitID', 'workerID')
    fieldsets = (
        (None, {'fields': ('assignment_id', 'data', 'phase', 'count')}),
    )
    """
    # Turn on to enable AMT HITs
    list_filter = (HITStatusFilter,)
    """

    def start_time(self, obj):
        return fromisoformat(obj.data['startTime'])
    start_time.admin_order_field = 'data__startTime'

    def worker_id(self, obj):
        return obj.data.get('workerId')
    worker_id.admin_order_field = 'data__workerId'

    def questions(self, obj):
        return format_html('<br>'.join("<a href={}>{}</a>".format(
            reverse('admin:{}_{}_change'.format(img._meta.app_label, img._meta.model_name),
            args=(img.id,)),
        img.text) for img in obj.questions.all()))
    def work_time(self, obj):
        end = obj.data.get('endTime')
        start = obj.data.get('startTime')
        if end and start:
            mins, secs = divmod((fromisoformat(end) - fromisoformat(start)).seconds, 60)
            return f"{mins} mins {secs} secs"
        else:
            return None

    """
    # Turn on to enable AMT HITs
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.registerAutoField('Feedback', 'RequesterFeedback')
    """
    def has_add_permission(self, request):
        return False

    def get_form(self, request, obj=None, **kwargs):
        defaults = {}
        if obj is None:
            defaults['form'] = HITCreationForm
            """
            # Turn on to enable AMT HITs
            if not messages.get_messages(request):
                hits = mturk.list_hits()['HITs']
                hitCounts = ((name, sum(1 for hit in hits if hit['HITStatus'] == 'Unassignable' and hit['Description'] == desc)) for name, desc in hitDescriptions.items())
                hitCountHuman = [f"{str(count)} {name}" for name, count in hitCounts if count]
                if hitCountHuman:
                    if len(hitCountHuman) == 1:
                        hitsHuman = hitCountHuman[0]
                    elif len(hitCountHuman) == 2:
                        hitsHuman = f"{hitCountHuman[0]} and {hitCountHuman[1]}"
                    else:
                        hitsHuman = f"{', '.join(hitCountHuman[:-1])}, and {hitCountHuman[-1]}"
                    self.message_user(request, f"There are {hitsHuman} HITs that are still in progress. Only proceed if you want more.", messages.WARNING)
            """
        else:
            self.fieldsets = (
                (None, {'fields': ('assignment_id', 'data')}),
            )
            self.readonly_fields = ('assignment_id',) # 'hitID', 'workerID'
        defaults.update(kwargs)
        return super().get_form(request, obj, **defaults)

    def status(self, obj):
        return None
    """
    # Turn on to enable AMT HITs
    def changelist_view(self, request, extra_context=None):
        # Update assignment statuses of any new assignments
        try:
            self.assignments = {
                assignment['AssignmentId']: assignment
                for hit in mturk.list_hits()['HITs']
                for assignment in mturk.list_assignments_for_hit(HITId=hit['HITId']).get('Assignments', [])
            }
            for obj in HIT.objects.filter(Q(data__has_key='hitId') &~ Q(data__has_key='status')):
                obj.data['status'] = self.assignments.get(obj.assignment_id, {}).get('AssignmentStatus', '')
                obj.save()
        except Exception as e:
            self.message_user(request, f'Unable to retrieve assignment statuses. {e}', messages.ERROR)
        return super().changelist_view(request, extra_context=None)
    def status(self, obj):
        if '__' in obj.assignment_id:
            return None
        if 'hitId' not in obj.data:
            return None
        #try:
        #    assignmentStatus = mturk.get_assignment(AssignmentId=obj.assignment_id).get('Assignment', {}).get('AssignmentStatus', '')
        #    obj.data['status'] = assignmentStatus
        #    obj.save()
        #except Exception as e:
        #    return None
        assignmentStatus = self.assignments.get(obj.assignment_id, {}).get('AssignmentStatus', '')
        # Update assignment statuses of any old assignments
        if 'status' in obj.data:
            if assignmentStatus == '':
                assignmentStatus = obj.data['status']
            elif obj.data['status'] != assignmentStatus:
                obj.data['status'] = assignmentStatus
                obj.save()
        else:
            obj.data['status'] = assignmentStatus
            obj.save()
        if assignmentStatus is '':
            return None
        icon_url = static('admin/img/icon-%s.svg' % {'Rejected': 'no', 'Approved': 'yes', 'Submitted': 'unknown'}[assignmentStatus])
        return format_html('<img src="{}" alt="{}">', icon_url, assignmentStatus)
    status.short_description = 'Assignment Status'
    """

    def phase(self, obj):
        return first(obj.data.get('roundnums', []), None)
    phase.admin_order_field = 'data__roundnums'

    """
    # Turn on to enable AMT HITs
    def registerAutoField(self, humanFieldName, fieldName):
        def field(obj):
            return self.assignments.get(obj.assignment_id, {}).get(fieldName, '')
        field.short_description = humanFieldName
        if not hasattr(self, fieldName):
            setattr(self, fieldName, field)
        self.list_display.append(fieldName)

    #self.list_display.remove(fieldName)

    def approve(self, request, queryset):
        for obj in queryset:
            try:
                mturk.approve_assignment(
                    AssignmentId=obj.assignment_id,
                    OverrideRejection=True
                )
            except Exception as e:
                self.message_user(request, f'Unable to approve the assignment {obj.assignment_id}. {e}', messages.ERROR)
    approve.short_description = 'Approve selected hits'
    def bonus(self, request, queryset):
        if request.POST.get('post'):
            form = HITApprovalForm(request.POST)
            if form.is_valid():
                for obj in queryset:
                    try:
                        #print('successful fake acceptance of ' + str(obj))
                        if form.cleaned_data['reason']:
                            mturk.approve_assignment(
                                AssignmentId=obj.assignment_id,
                                RequesterFeedback=form.cleaned_data['reason'],
                                OverrideRejection=True
                            )
                        else:
                            mturk.approve_assignment(
                                AssignmentId=obj.assignment_id,
                                OverrideRejection=True
                            )
                        obj.data['__cached_status'] = 'Approved'
                        obj.save()
                        if form.cleaned_data['bonus']:
                            #print('successful fake bonus of '+str(form.cleaned_data['bonus']))
                            mturk.send_bonus(
                                WorkerId=obj.workerID,
                                BonusAmount=str(form.cleaned_data['bonus']),
                                AssignmentId=obj.assignment_id,
                                Reason=form.cleaned_data['reason'],
                                UniqueRequestToken=obj.assignment_id
                            )
                    except Exception as e:
                        self.message_user(request, f'Unable to approve the assignment {obj.assignment_id}. {e}', messages.ERROR)
                return None
        else:
            form = HITApprovalForm()
        return render(request, 'admin/users/hit/hit_form.html', {
            'items': queryset.order_by('pk'),
            'form': form,
            'title': 'Approve selected hits',
            'action': 'bonus',
            'button': 'Approve',
        })
    bonus.short_description = 'Approve selected hits and send bonus'

    def reject(self, request, queryset):
        if request.POST.get('post'):
            form = HITRejectionForm(request.POST)
            if form.is_valid():
                for obj in queryset:
                    try:
                        #print('successful fake rejection of ' + str(obj))
                        mturk.reject_assignment(
                            AssignmentId=obj.assignment_id,
                            RequesterFeedback=form.cleaned_data['reason']
                        )
                        obj.data['__cached_status'] = 'Rejected'
                        obj.save()
                    except Exception as e:
                        self.message_user(request, f'Unable to reject the assignment {obj.assignment_id}. {e}', messages.ERROR)
                return None
        else:
            form = HITRejectionForm()
        return render(request, 'admin/users/hit/hit_form.html', {
            'items': queryset.order_by('pk'),
            'form': form,
            'title': 'Reject selected hits',
            'action': 'reject',
            'button': 'Reject',
        })
    reject.short_description = 'Reject selected hits'
    actions=[approve, bonus, reject]
    """

class AnswerFrameIDFilter(admin.SimpleListFilter):
    title = 'Task'
    parameter_name = 'task'
    def lookups(self, request, model_admin):
        return [('Clustered', 'Clustered'), ('Random', 'Random')]
    def queryset(self, request, queryset):
        val = self.value()
        if val is None:
            return queryset
        if val == 'Clustered':
            return queryset.filter(~Q(question__frameID=-1))
        return queryset.filter(question__frameID=-1)

class FrameIDFilter(admin.SimpleListFilter):
    title = 'Task'
    parameter_name = 'task'
    def lookups(self, request, model_admin):
        return [('Clustered', 'Clustered'), ('Random', 'Random')]
    def queryset(self, request, queryset):
        val = self.value()
        if val is None:
            return queryset
        if val == 'Clustered':
            return queryset.filter(~Q(frameID=-1))
        return queryset.filter(frameID=-1)

class AnswerAdmin(admin.ModelAdmin, ExportCSVMixin):
    actions = ['export_csv']
    export_filename = 'phase1-answers.csv'
    export_field_names = ['id','text','isFinal','question','hit']
    list_filter = ('isFinal', ('hit', admin.RelatedOnlyFieldListFilter), (AnswerFrameIDFilter))
    list_display = ('text', 'question', 'hit')

class AnswerInline(admin.TabularInline):
    model = Answer
    extra = 0

class QuestionAdmin(admin.ModelAdmin, ExportCSVMixin):
    actions = ['export_csv']
    export_filename = 'phase1-questions.csv'
    export_field_names = ['id','text','isFinal','hit','mergeParent']

    inlines = [
        AnswerInline,
    ]

    readonly_fields = ('image_id', 'merge_parent', 'merge_children')
    exclude = ('imageID', 'mergeParent')
    def image_id(self, obj):
        id = obj.imageID
        imgs = ImageModel.objects.filter(img__in=id) if id else []
        return format_html('<br>'.join("<a href={}>{}</a>".format(
            reverse('admin:{}_{}_change'.format(img._meta.app_label, img._meta.model_name),
            args=(img.pk,)),
        img.img) for img in imgs))
    image_id.short_description = 'Image ID'
    def merge_parent(self, obj):
        return format_html("<a href={}>{}</a>".format(
            reverse('admin:{}_{}_change'.format(obj._meta.app_label, obj._meta.model_name),
            args=(obj.mergeParent,)),
        Question.objects.get(id=obj.mergeParent)) if obj.mergeParent else "Question is final")
    def merge_children(self, obj):
        children = Question.objects.filter(mergeParent=obj.id)
        if children:
            return format_html(f'<li>{"</li><li>".join(self._merge_children(child) for child in children)}</li>')
        return format_html("No children")
    def _merge_children(self, obj):
        children = Question.objects.filter(mergeParent=obj.id)
        me = "<a href={}>{}</a>".format(
            reverse('admin:{}_{}_change'.format(obj._meta.app_label, obj._meta.model_name),
            args=(obj.id,)),
        obj)
        if children:
            return f'{me}<ul><li>{"</li><li>".join(self._merge_children(child) for child in children)}</li></ul>'
        return me
    list_filter = ('isFinal', ('hit', admin.RelatedOnlyFieldListFilter), FrameIDFilter)
    list_display = ('text', 'merge_children')


class ExperimentAdmin(admin.ModelAdmin):
    def exportData(self, request, queryset):
        if len(queryset) == 1:
            experiment = queryset[0]
            response = HttpResponse(ExperimentAdmin.exportOneExperiment(experiment), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment;filename=experiment-{experiment}.xlsx'
        else:
            with tempfile.SpooledTemporaryFile() as tmp:
                with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as archive:
                    for experiment in queryset:
                        archive.writestr(f'experiment-{experiment}.xlsx', ExperimentAdmin.exportOneExperiment(experiment))
                tmp.seek(0)
                response = HttpResponse(tmp.read(), content_type='application/x-zip-compressed')
                response['Content-Disposition'] = 'attachment;filename=experiments.zip'
        return response
    actions = [exportData]

    @classmethod
    def exportOneExperiment(cls, experiment):
        # Read data from database
        all = HIT.objects.count()
        submitted = HIT.objects.filter(data__status='Submitted').count()
        accepted = HIT.objects.filter(data__status='Accepted').count()
        rejected = HIT.objects.filter(data__status='Rejected').count()
        status = ';'.join([i for _, i in MigrationLoader(connection=connection, ignore_no_migrations=True).graph.leaf_nodes('users')])

        # Put data on first sheet
        book = Workbook() #write_only=True
        metasheet = book.active #book.create_sheet("Experiment Information")
        metasheet.title = "Experiment Information"
        metasheet.append(("Experiment ID", -1)) #experiment.id
        metasheet.append(("Start Date", -2)) #experiment.start
        metasheet.append(("End Date", -3 or "Ongoing")) #experiment.end
        metasheet.append(("Export Date", datetime.now()))
        metasheet.append((None, None))
        metasheet.append(("HITs Created", -4)) #experiment.hits
        metasheet.append(("HITs Attempted", all))
        metasheet.append(("HITs Completed", submitted + accepted + rejected))
        metasheet.append(("HITs Accepted", accepted))
        metasheet.append(("HITs Rejected", rejected))
        metasheet.append((None, None))
        metasheet.append(("Migration Status", status))
        metasheet.sheet_state = 'hidden'

        # Create the other sheets
        questions = cls.exportSheet(book,
            Question.objects.all().order_by('-isFinal', 'id'),
            'id','text','isFinal','hit','mergeParent',
            mergeParent=None
        )
        answers = cls.exportSheet(book,
            Answer.objects.all().order_by('-isFinal', 'id'),
            'id','text','isFinal','question','hit',
            question=questions,
        )
        attributes = cls.exportSheet(book,
            Attribute.objects.all().order_by('count'),
            'word','count','weight','question:answer.question','answer',
            question=questions,
            answer=answers,
        )

        return save_virtual_workbook(book)

    tableStyle = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    @classmethod
    def exportSheet(cls, book, qs, *cols, **links):
        name = qs.model._meta.verbose_name_plural.title()
        sheet = book.create_sheet(name)
        listqs = list(qs)
        lastCol = get_column_letter(len(cols))

        # Parse column names
        cols = cls.getFields(cols)

        # Add header
        sheet.freeze_panes = 'A2'
        sheet.append([col[0] for col in cols])

        # Import data
        for i, obj in enumerate(listqs, 2):
            for j, (colName, colGetter) in enumerate(cols, 1):
                val = colGetter(obj)
                link = links.get(colName)
                cell = sheet.cell(i, j, str(val) if isinstance(val, Model) else val)
                if link:
                    try:
                        rowId = link[1].index(val) + 2
                        cell.hyperlink = f"#{link[0]}!A{rowId}:{link[2]}{rowId}"
                    except ValueError:
                        pass

        # Create recursive links
        recursive_links = [k for k, v in links.items() if v is None]
        if recursive_links:
            for j, (colName, colGetter) in enumerate(cols, 1):
                if colName in recursive_links:
                    for i, obj in enumerate(listqs, 2):
                        val = colGetter(obj)
                        cell = sheet.cell(i, j)
                        try:
                            val = val if isinstance(val, qs.model) else qs.model.objects.get(pk=val)
                            rowId = listqs.index(val) + 2
                            cell.hyperlink = f"#{name}!A{rowId}:{lastCol}{rowId}"
                        except ValueError:
                            pass
                        except qs.model.DoesNotExist:
                            pass

        # Create table object
        #sheet.add_table(Table(ref=sheet.calculate_dimension(), displayName=name, tableStyleInfo=cls.tableStyle))

        # Create model data for linking
        return (name, listqs, lastCol)

    @classmethod
    def getFields(cls, names):
        fields = []
        for name in names:
            #if '%' in name: # eval is scary. Please be friendly.
            #    header, field = name.split(';', 2)
            #    fields.append((header, eval(field)))
            if ':' in name:
                header, field = name.split(':')
                fields.append((header, operator.attrgetter(field)))
            else:
                fields.append((name, operator.attrgetter(name)))
        return fields

    def has_add_permission(self, request):
        return False
    def has_change_permission(self, request, obj=None):
        return False

    def pre_validate_step1(self):
        countImgs = ImageModel.objects.filter(img__contains=settings.KEYRING).count()
        if countImgs % 3 == 1 or countImgs % 4 == 1:
            raise forms.ValidationError('The number of images in a dataset is not allowed to be 1 more than a multiple of 3 or 4 for Steps 1 and 2 to run.')

    def pre_validate_step2(self):
        countQs = Question.objects.filter(isFinal=True)
        minQs = settings.NUMROUNDS[phase01b.__name__]
        if countQs < 1:
            raise forms.ValidationError('The number of images in a dataset is not allowed to be 1 more than a multiple of 3 or 4 for Steps 1 and 2 to run.')
        elif countQs < minQs:
            self.message_user(request, f"There are {countQs} questions that were created in Step 1. It is recommended that you get at least {minQs} before continuing.", messages.WARNING)

    def pre_validate_step3(self):
        countQs = Question.objects.filter(isFinal=True)
        countAs = Attribute.objects.count()
        if countQs != countAs:
            raise forms.ValidationError('The answer reducing NLP script was not run correctly. Please delete all Attribute objects and rerun the script.')
        elif countAs < 1:
            raise forms.ValidationError('Step 1 was not run yet. Please finish Steps 1 and 2 before going on to Step 3.')

admin.site.register(CustomUser, CustomUserAdmin)
# admin.site.register(Zipfile)
admin.site.register(Phase, PhaseAdmin)

admin.site.register(Attribute, AttributeAdmin)
admin.site.register(ImageModel, ImageModelAdmin)

admin.site.register(Phase01_instruction)
admin.site.register(Phase02_instruction)
admin.site.register(Phase03_instruction)
admin.site.register(TextInstruction)
admin.site.register(Question, QuestionAdmin)
admin.site.register(Answer, AnswerAdmin)
admin.site.register(Session, SessionAdmin)
admin.site.register(HIT, HITAdmin)
